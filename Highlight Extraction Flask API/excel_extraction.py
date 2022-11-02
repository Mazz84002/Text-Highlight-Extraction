import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def excel_extraction(path):
    wb = load_workbook(
        filename=path)
    list_of_sheets = wb.sheetnames
    # print(list_of_sheets)

    highlights = {}

    def annotation_checker(cell_obj):
        color_in_hex = cell_obj.fill.start_color.index
        # print(color_in_hex)
        color_list = ['FF92D050', 'FFFFFF00']  # default colour-codes used for highligting in excel
        if color_in_hex in color_list:
            return True
        return False

    # iterate over all of the sheets
    for sheet in list_of_sheets:
        # check the number of rows and colums
        temp_list = []
        sh = wb[sheet]
        for i in range(1, sh.max_row + 1):  # iterate over all the rows
            for j in range(1, sh.max_column + 1):
                cell_obj = sh.cell(row=i, column=j)
                # print(cell_obj.fill)
                if annotation_checker(cell_obj):  # annotation checker
                    temp_list.append((cell_obj.value, str(cell_obj.row) + cell_obj.column_letter))
        #print(temp_list)
        if temp_list != []:
            highlights[sheet] = temp_list
    return highlights